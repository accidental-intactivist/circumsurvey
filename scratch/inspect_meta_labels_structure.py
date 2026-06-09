from openpyxl import load_workbook
wb = load_workbook('data/responses.xlsx', data_only=True)
ws = wb['Meta Labels']
rows = list(ws.iter_rows(values_only=True))
print("Meta Labels Dimensions: Rows =", len(rows), "Cols =", len(rows[0]) if rows else 0)
print("Header:", rows[0])
for row in rows[1:15]:
    if any(row):
        print(row)
