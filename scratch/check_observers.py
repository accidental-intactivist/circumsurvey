from openpyxl import load_workbook
import sys

wb = load_workbook(sys.argv[1], data_only=True)
ws = wb['Form Responses 1']
header = [str(c).strip() if c else '' for c in next(ws.iter_rows(values_only=True, min_row=1, max_row=1))]

# Find rows where col 65 is empty (the 38 observers)
for row_idx, row in enumerate(ws.iter_rows(values_only=True, min_row=2), start=2):
    pathway_val = row[65] if len(row) > 65 else None
    if pathway_val is not None:
        continue
    # This is an "empty pathway" row — scan nearby columns for observer clues
    print(f"\n--- Row {row_idx} ---")
    for col_idx in range(60, min(75, len(row))):
        val = row[col_idx]
        if val:
            h = header[col_idx] if col_idx < len(header) else f"col{col_idx}"
            print(f"  col {col_idx}: {repr(str(val)[:100])}  (header: {h[:60]})")
    if row_idx > 5:  # Just check first few
        break
