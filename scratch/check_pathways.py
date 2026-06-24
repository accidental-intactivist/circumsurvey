from openpyxl import load_workbook
import sys

wb = load_workbook(sys.argv[1], data_only=True)
ws = wb['Form Responses 1']
vals = {}
for row in ws.iter_rows(values_only=True, min_row=2):
    v = str(row[65]).strip() if row[65] else 'EMPTY'
    vals[v] = vals.get(v, 0) + 1

for v in sorted(vals.keys()):
    print(f"{vals[v]:4d}  {repr(v)}")
