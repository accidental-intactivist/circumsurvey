#!/usr/bin/env python3
"""
CircumSurvey — build_questions_metadata.py

Scans the Form Responses xlsx to build the `questions` metadata table,
which drives the Explore and Data subdomain UIs without hardcoded JS.

For each of the ~373 columns, this script determines:

  id          — stable slug (meta_tag from Meta Tag Mapping, or q### fallback)
  prompt      — the full question wording from the header row
  section     — Demographics | Family | Religion | Experience | Health |
                Identity | Attitudes | Partner | Parenting | Trans | Intersex
  pathway     — all | intact | circumcised | restoring | observer | trans | intersex
                (inferred from emoji prefix in the header)
  type        — scale_1_5 | single_select | multi_select | open_text
  opts_json   — JSON array of distinct non-empty values (null for open_text)
  tier        — 1 (curated, already on findings site), 2 (browseable deep-dive),
                3 (long tail, indexed but not featured)
  col_idx     — 0-based original column index

Outputs:
  data/questions_seed.sql — INSERT statements for the questions table
  data/questions_manifest.json — JSON version for review/auditing

Usage:
  python3 scripts/build_questions_metadata.py \\
      --xlsx data/responses.xlsx \\
      --out-sql data/questions_seed.sql \\
      --out-json data/questions_manifest.json
"""

import argparse
import json
import re
from collections import Counter
from pathlib import Path

# ─────────────────────────────────────────────────────────────
# Pathway inference from emoji prefixes in column headers
# ─────────────────────────────────────────────────────────────
EMOJI_TO_PATHWAY = {
    "🟢": "intact",
    "🔵": "circumcised",
    "🟣": "restoring",
    "🟠": "observer",
    "🔴": "trans",
    "⚪": "intersex",
    "🟡": "all",      # religion sections use yellow for Christian-specific
}

# ─────────────────────────────────────────────────────────────
# Section inference from slug prefix
# ─────────────────────────────────────────────────────────────
SLUG_SECTION_MAP = [
    ("meta_",            "Meta"),
    ("pathway_",         "Pathway Routing"),
    ("demo_",            "Demographics"),
    ("family_",          "Family"),
    ("religion_",        "Religion"),
    ("culture_",         "Culture & Attitudes"),
    ("final_",           "Culture & Attitudes"),
    ("contact_",         "Follow-up"),
    ("exp_sex_",         "Sexual Experience"),
    ("exp_appearance",   "Appearance"),
    ("exp_orgasm",       "Sexual Experience"),
    ("exp_pleasure",     "Sexual Experience"),
    ("exp_lubrication",  "Sexual Experience"),
    ("exp_pride",        "Pride & Regret"),
    ("exp_sensitivity",  "Sexual Experience"),
    ("exp_",             "Experience"),
    ("health_",          "Health"),
    ("identity_",        "Identity"),
    ("attitude_",        "Culture & Attitudes"),
    # Pathway-scoped sections
    ("intact_",          "Intact Pathway"),
    ("circ_",            "Circumcised Pathway"),
    ("restore_",         "Restoring Pathway"),
    ("observe_",         "Observer Pathway"),
    ("trans_",           "Trans Pathway"),
    ("intersex_",        "Intersex Pathway"),
    ("q",                "Uncategorized"),
]

# ─────────────────────────────────────────────────────────────
# Tier 1: the 19 curated questions featured on findings.circumsurvey.online
# ─────────────────────────────────────────────────────────────
TIER_1_SLUGS = {
    # Sexual experience rating grid (6 curated)
    "exp_sex_rating_orgasm_intensity",
    "exp_sex_rating_orgasm_duration",
    "exp_sex_rating_ease_of_orgasm",
    "exp_sex_rating_sensitivity_light_touch",
    "exp_sex_rating_pleasure_mobile_skin",
    "exp_sex_rating_variety_of_sensation",
    # Appearance & pride (featured)
    "exp_appearance_feeling",
    "exp_pride_satisfaction_rating",
    # Cultural / attitudinal flagships
    "final_aesthetic_preference",
    "final_social_norm_perception",
    "culture_assoc_more_aesthetic",
    # Pathway-anchored flagships
    "intact_regret_feeling",
    "circ_restoration_awareness",
    "restore_impact_rating_aesthetics",
    "circ_presentation_by_medical",
    # Family & autonomy
    "family_father_status",
    "circ_adult_consent_status",
    # Cross-pathway flagship
    "intact_pressure_to_circ",
}


def normalize_header(s):
    if not s:
        return ""
    return ' '.join(str(s).strip().split())


def strip_emoji(s):
    """Strip emoji and return (clean_text, detected_pathway)."""
    if not s:
        return "", None
    s = str(s)
    pathway = None
    for emoji, pw in EMOJI_TO_PATHWAY.items():
        if emoji in s:
            pathway = pw
            s = s.replace(emoji, "")
            break  # only take the first emoji seen
    # Also strip any remaining emoji-range chars
    s = re.sub(
        r'[\U0001F300-\U0001FAFF\u2600-\u27BF\u2B00-\u2BFF\uFE0F\u20E3]+',
        '', s
    )
    s = ' '.join(s.split()).strip(' \t\n\r.:;-–—')
    return s, pathway


def section_from_slug(slug):
    for prefix, section in SLUG_SECTION_MAP:
        if slug.startswith(prefix):
            return section
    return "Uncategorized"


def split_outside_paren(s):
    """Splits a string by commas, but only if the comma is outside of parentheses."""
    parts = []
    current = []
    paren_depth = 0
    i = 0
    while i < len(s):
        c = s[i]
        if c == '(':
            paren_depth += 1
            current.append(c)
        elif c == ')':
            paren_depth -= 1
            current.append(c)
        elif c == ',' and paren_depth == 0:
            parts.append("".join(current).strip())
            current = []
            if i + 1 < len(s) and s[i+1] == ' ':
                i += 1
        else:
            current.append(c)
        i += 1
    if current:
        parts.append("".join(current).strip())
    return [p for p in parts if p]


FORCED_QUESTION_CHOICES = {
    "circ_parents_influences": (
        "multi_select",
        [
            'Direct Medical Authority: Specific advice from their obstetrician, pediatrician, or the hospital staff where you were born. (The voice of a trusted professional).',
            'Institutional Medical Norm: The general belief that "this is just the standard procedure" at hospitals, making it a routine, unquestioned part of postpartum care. (The power of the system).',
            'Family Tradition/Pressure: The expectation from grandparents or other influential relatives; the idea that "it\'s what the men in our family have always done." (The power of heritage).',
            'Paternal Influence (The "Like Father" Factor): A specific desire for you to genitally match your father or other male family members. (The power of conformity within the family unit).',
            'Peer & Social Pressure (The "Fitting In" Factor): The widespread belief that "everyone else is doing it," and a desire for you to avoid being seen as different in locker rooms or later in life. (The power of broad social conformity).',
            'Religious Mandate/Tradition: A direct requirement or strong cultural expectation stemming from their religious beliefs and community. (The power of faith).',
            'Prevailing Health & Hygiene Beliefs: General societal "knowledge" at the time—gleaned from various sources—that presented circumcision as being fundamentally cleaner, healthier, or preventive of future diseases. (The power of a cultural \'truth\').',
            'Popular Media & Parenting "Experts": Information or attitudes transmitted through widely-read parenting books (like Dr. Spock in earlier eras), magazines, or other mainstream media. (The power of mass media).',
            'Lack of Counter-Information: Crucially, the simple absence of any strong, readily available information presenting the benefits of intactness or the risks of circumcision, making it an easy default choice. (The power of silence).',
            'Aesthetic Preference: A belief, possibly influenced by media or pornography, that the circumcised penis is visually "neater," "more attractive," or the "correct" look. (The power of aesthetics).',
            'I have absolutely no idea what influenced them.',
            'Prevailing Moral Beliefs about Sexuality (e.g., concern over masturbation):'
        ]
    ),
    "family_mother_profession": (
        "multi_select",
        [
            'Stay-at-Home Parent / Homemaker',
            'Clerical / Administrative Support (e.g., Secretary, Office Clerk)',
            'Education (e.g., K-12 Teacher, Professor, School Administrator)',
            'Business/Finance/Management',
            'Healthcare/Medicine (e.g., Doctor, Nurse, Therapist, Technician)',
            'Law/Government/Public Service/Civil Servant',
            'Science/Research/Academia',
            'Factory / Manufacturing / General Labor',
            'Arts/Humanities/Entertainment (e.g., Artist, Writer, Musician)',
            'Skilled Trades (e.g., Electrician, Carpenter, Mechanic, Cosmetologist)',
            'Personal Care / Service',
            'Retail / Customer Service / Hospitality',
            'Military Service',
            'I\'m not sure'
        ]
    ),
    "family_father_profession": (
        "multi_select",
        [
            'Skilled Trades (e.g., Electrician, Carpenter, Mechanic, Cosmetologist)',
            'Business/Finance/Management',
            'Factory / Manufacturing / General Labor',
            'Law/Government/Public Service',
            'Healthcare/Medicine (e.g., Doctor, Nurse, Therapist, Technician)',
            'Retail / Customer Service / Hospitality',
            'Science/Research/Academia',
            'Education (e.g., K-12 Teacher, Professor, School Administrator)',
            'Arts/Humanities/Entertainment',
            'Military Service',
            'Clerical / Administrative Support (e.g., Secretary, Office Clerk)',
            'Stay-at-Home Parent / Homemaker',
            'Personal Care / Service',
            'I\'m not sure'
        ]
    ),
    "intact_parents_info_sources": (
        "multi_select",
        [
            'Advice from a specific group or community (e.g., La Leche League, a natural health co-op)',
            'A specific medical statement or article (e.g., the 1971 AAP statement disavowing medical need)',
            'Family history or cultural background from a non-circumcising country/culture',
            'Influence from a foreign-born, foreign-trained, or older doctor/nurse who was skeptical of the practice',
            'Don\'t Know / Unsure'
        ]
    ),
    "intact_parents_traits_values": (
        "multi_select",
        [
            'Highly valued scientific evidence and did their own research',
            'Fiscally conscious (questioned unnecessary medical costs)',
            'Strongly principled about consent and bodily autonomy in general',
            'Valued natural, holistic, or low-intervention approaches to health',
            'Identified as non-conformist or counter-cultural',
            'Generally skeptical of medical authority or "doctor knows best"',
            'None of the above feel particularly accurate',
            'Unsure'
        ]
    ),
    "circ_adult_motivation_details": (
        "multi_select",
        [
            'Aesthetic Preference: I (or my partner/parents) preferred the appearance of a circumcised penis.',
            'Perceived Hygiene Benefits: A belief that it would be cleaner or easier to care for.',
            'Social Conformity: A desire to "look like" peers or fit a cultural norm.',
            'Sensation: A belief or hope that it would change or improve sexual sensation (e.g., help with premature ejaculation).',
            'Medical Diagnosis: Phimosis (inability to retract the foreskin)',
            'Medical Diagnosis: Recurring Balanitis/Infections',
            'Medical Diagnosis: Other specific condition',
            'Religious Conversion or Affirmation'
        ]
    ),
    "observe_curious_shaping_factors": (
        "multi_select",
        [
            'Personal experiences with intimate partners',
            'Independent research I\'ve done myself (intactivist sites, documentaries, etc.)',
            'Conversations with male friends or partners about their experiences',
            'Feminist principles of bodily autonomy and consent',
            'Information or advice from doctors or medical professionals',
            'Media portrayals (movies, TV, internet)',
            'Conversations with friends',
            'Religious or cultural teachings',
            'The norms in my family or the community I grew up in'
        ]
    ),
    "observe_parent_intact_factors": (
        "multi_select",
        [
            'Ethical beliefs about bodily autonomy and a child\'s right to consent',
            'Information about the functions and sensitivity of the foreskin',
            'Lack of a clear medical necessity for the procedure',
            'Concerns about surgical risks and potential complications',
            'Independent research (documentaries, articles, online communities)',
            'My partner\'s preference was to keep him intact',
            'The medical benefit claims did not make sense.'
        ]
    ),
    "intact_parents_neg_catalyst": (
        "multi_select",
        [
            'Yes, I believe one of my parents (e.g., my father) had a negative personal experience with their own circumcision.',
            'No',
            'Unsure / Don\'t Know'
        ]
    ),
    "demo_ethnicity": (
        "multi_select",
        [
            'Asian / Asian American (e.g., East Asian, South Asian, Southeast Asian)',
            'Black / African American / African / Afro-Caribbean',
            'Native American / Alaska Native / Indigenous / First Nations',
            'White / Caucasian / European American',
            'Hispanic / Latino / Latina / Latinx',
            'Native Hawaiian / Other Pacific Islander',
            'Middle Eastern / North African (MENA)',
            'Multiracial / Biracial',
            'Prefer not to say'
        ]
    ),
    "demo_race_ethnicity": (
        "multi_select",
        [
            'Asian / Asian American (e.g., East Asian, South Asian, Southeast Asian)',
            'Black / African American / African / Afro-Caribbean',
            'Native American / Alaska Native / Indigenous / First Nations',
            'White / Caucasian / European American',
            'Hispanic / Latino / Latina / Latinx',
            'Native Hawaiian / Other Pacific Islander',
            'Middle Eastern / North African (MENA)',
            'Multiracial / Biracial',
            'Prefer not to say'
        ]
    ),
    "restore_techniques_used": (
        "multi_select",
        [
            'Manual tugging (Andre\'s method, etc)',
            'T-Tape',
            'O-rings / retaining cones',
            'Dual-tension devices (DTR, Mantis, etc)',
            'Air inflation devices (Foreskinned Air, HyperRestore, etc)',
            'Weights (PUD, stealth retainers with weights, etc)',
            'Surgical restoration / Foregen clinical trials',
            'I haven\'t started yet'
        ]
    ),
    "observe_advocate_future_focus": (
        "multi_select",
        [
            'Legal challenges and lawsuits (like the Equal Protection cases)',
            'Legislative action (e.g., defunding Medicaid for RIC)',
            'Direct outreach and education for expectant parents',
            'Reforming medical school curricula and hospital protocols',
            'High-visibility public protests and awareness campaigns',
            'Creating high-quality media (documentaries, articles)',
            'Supporting foreskin restoration and regeneration research',
            'Building broader coalitions with other human rights groups'
        ]
    ),
    "religion_jewish_brit_milah_view": (
        "single_select",
        [
            "As a non-negotiable religious commandment (mitzvah) central to Jewish identity.",
            "As an important tradition and cultural marker, though perhaps with some flexibility in interpretation or practice.",
            "As a practice open to discussion, with varying opinions on its necessity or form in modern times.",
            "As a practice that some in my community/family questioned or chose alternatives to.",
            "Primarily as a medical/hygienic procedure that also fulfilled a religious/cultural role.",
            "Not a significant topic of discussion or emphasis in my specific context."
        ]
    )
}


def infer_type_and_opts(slug, values):
    """
    Given the list of non-empty response values for a column, infer:
      - type: scale_1_5 | single_select | multi_select | open_text
      - opts: list of distinct option labels (or None for open_text)

    Heuristics:
      - If ≥95% of values are 1-5 integers or 'N-label' patterns → scale_1_5
      - If ≥30% of values contain commas and distinct count > 2 → multi_select
      - If distinct count ≤ 25 and max length ≤ 200 → single_select
      - Otherwise → open_text
    """
    if slug in FORCED_QUESTION_CHOICES:
        return FORCED_QUESTION_CHOICES[slug]

    if not values:
        return "open_text", None

    n = len(values)
    distinct = Counter(values)

    # Scale detection: values starting with "1", "2", ..., "5" followed by
    # punctuation/space OR just bare integers
    scale_pat = re.compile(r"^\s*[1-5](\.0)?(\s*[-–—:]|\s|$)")
    scale_count = sum(1 for v in values if scale_pat.match(str(v)))
    if scale_count / n >= 0.95 and len(distinct) <= 10:
        # Extract the distinct scale labels for display
        return "scale_1_5", sorted(set(values), key=lambda v: str(v)[:2])

    # Multi-select detection: lots of comma-containing values
    comma_count = sum(1 for v in values if ',' in str(v))
    if comma_count / n >= 0.30 and len(distinct) > 2:
        # Build option list by splitting on commas (outside paren) and deduping
        opts = set()
        for v in values:
            parts = split_outside_paren(str(v))
            opts.update(parts)
        # Guard against runaway splits (commas inside free text)
        if len(opts) <= 50:
            return "multi_select", sorted(opts)
        # Too many distinct parts → probably free text
        return "open_text", None

    # Single-select: bounded cardinality + short strings
    max_len = max((len(str(v)) for v in values), default=0)
    if len(distinct) <= 25 and max_len <= 200:
        return "single_select", sorted(distinct.keys())

    return "open_text", None


def infer_tier(slug, header, pathway, qtype):
    """Assign tier 1/2/3 for UI prioritization."""
    if slug in TIER_1_SLUGS:
        return 1
    # Pathway-scoped depth questions = tier 2 (for Explore site)
    if pathway and pathway != "all" and qtype != "open_text":
        return 2
    # Non-pathway single/multi/scale questions = tier 2
    if qtype in ("scale_1_5", "single_select", "multi_select"):
        return 2
    # Free text + q### fallback slugs = tier 3
    return 3


def load_form_data(xlsx_path):
    """Returns (header_row, data_rows, col_to_meta dict)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)

    # Build header -> meta_tag map from Meta Tag Mapping sheet
    header_to_meta = {}
    if "Meta Tag Mapping" in wb.sheetnames:
        ws_meta = wb["Meta Tag Mapping"]
        for row in ws_meta.iter_rows(values_only=True):
            if row and row[0] and row[1]:
                h = strip_emoji(normalize_header(row[0]))[0]
                m = str(row[1]).strip()
                if h and m:
                    header_to_meta[h] = m
        header_to_meta_lower = {k.lower(): v for k, v in header_to_meta.items()}
    else:
        header_to_meta_lower = {}

    ws = wb["Form Responses 1"]
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[0]
    data = [r for r in all_rows[1:] if r and r[0]]

    # Map each col → meta_tag slug
    col_to_meta = {}
    for i, h in enumerate(header):
        if not h:
            continue
        clean, _ = strip_emoji(normalize_header(h))
        if clean in header_to_meta:
            col_to_meta[i] = header_to_meta[clean]
        elif clean.lower() in header_to_meta_lower:
            col_to_meta[i] = header_to_meta_lower[clean.lower()]
        else:
            # Substring fallback
            prefix = clean[:40].lower()
            for mh, ms in header_to_meta_lower.items():
                if prefix and (mh.startswith(prefix) or prefix in mh):
                    col_to_meta[i] = ms
                    break

    # Manual fills for religion sub-questions and pathway column
    religion_manual = {
        24: "religion_christian_denomination", 25: "religion_christian_circ_view",
        26: "religion_christian_theology_basis", 27: "religion_christian_comments",
        28: "religion_jewish_denomination", 29: "religion_jewish_brit_milah_view",
        30: "religion_jewish_theology_awareness", 31: "religion_jewish_theology_reasons",
        32: "religion_jewish_identity_importance", 33: "religion_jewish_alt_interpretations",
        34: "religion_jewish_alt_thoughts", 35: "religion_jewish_diversity_room",
        36: "religion_jewish_brit_shalom", 37: "religion_islamic_madhhab",
        38: "religion_islamic_khitan_view", 39: "religion_islamic_religious_awareness",
        40: "religion_islamic_religious_reasons", 41: "religion_islamic_fard_or_sunnah",
        42: "religion_islamic_identity_importance", 43: "religion_islamic_alt_interpretations",
        44: "religion_islamic_alt_thoughts", 45: "religion_islamic_diversity_room",
        46: "religion_islamic_intactness_reconcile",
    }
    for k, v in religion_manual.items():
        col_to_meta[k] = v

    col_to_meta[65] = "pathway_state"

    for i in range(len(header)):
        if header[i] is not None and i not in col_to_meta:
            col_to_meta[i] = f"q{i:03d}"

    return header, data, col_to_meta


def build_questions(header, data, col_to_meta):
    """For each column, build a question metadata dict."""
    questions = []

    for col_idx in range(len(header)):
        raw_header = header[col_idx]
        if not raw_header:
            continue

        slug = col_to_meta.get(col_idx)
        if not slug:
            continue

        prompt_clean, pathway = strip_emoji(normalize_header(raw_header))
        if pathway is None:
            pathway = "all"

        # Collect non-empty values for this column
        values = []
        for row in data:
            if col_idx < len(row):
                v = row[col_idx]
                if v is None:
                    continue
                v_str = str(v).strip()
                if v_str:
                    values.append(v_str)

        qtype, opts = infer_type_and_opts(slug, values)
        section = section_from_slug(slug)
        tier = infer_tier(slug, prompt_clean, pathway, qtype)

        # For very long prompts, also extract a subtitle (text after first sentence)
        subtitle = None
        if len(prompt_clean) > 120:
            # Try splitting on first question-mark or colon
            m = re.search(r'[?:]', prompt_clean)
            if m:
                subtitle = prompt_clean[m.end():].strip()
                prompt_clean = prompt_clean[:m.end()].strip()

        questions.append({
            "id": slug,
            "section": section,
            "pathway": pathway,
            "prompt": prompt_clean,
            "subtitle": subtitle,
            "type": qtype,
            "opts": opts,
            "tier": tier,
            "col_idx": col_idx,
            "n_responses": len(values),
        })

    # De-duplicate: if two columns share the same slug (shouldn't happen but
    # guard for it), keep the first and drop the rest
    seen = {}
    deduped = []
    for q in questions:
        if q["id"] not in seen:
            seen[q["id"]] = True
            deduped.append(q)
    return deduped


def sql_escape(s):
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def write_sql(questions, out_path):
    lines = [
        "-- ═══════════════════════════════════════════════════════════════",
        "-- CircumSurvey — questions metadata seed",
        "-- Generated by scripts/build_questions_metadata.py",
        f"-- {len(questions)} questions",
        "-- Run: npx wrangler d1 execute circumsurvey --remote --file data/questions_seed.sql",
        "-- ═══════════════════════════════════════════════════════════════",
        "",
        "DELETE FROM questions;",
        "",
    ]
    for q in questions:
        opts_json = sql_escape(json.dumps(q["opts"], ensure_ascii=False)) if q["opts"] else "NULL"
        lines.append(
            "INSERT INTO questions (id, section, pathway, prompt, subtitle, type, opts_json, tier, col_idx) VALUES ("
            f"{sql_escape(q['id'])}, "
            f"{sql_escape(q['section'])}, "
            f"{sql_escape(q['pathway'])}, "
            f"{sql_escape(q['prompt'])}, "
            f"{sql_escape(q['subtitle'])}, "
            f"{sql_escape(q['type'])}, "
            f"{opts_json}, "
            f"{q['tier']}, "
            f"{q['col_idx']}"
            ");"
        )
    lines.append("")
    lines.append(f"-- {len(questions)} rows inserted")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")


def write_json(questions, out_path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(questions, indent=2, ensure_ascii=False), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out-sql", required=True)
    ap.add_argument("--out-json", required=True)
    args = ap.parse_args()

    header, data, col_to_meta = load_form_data(Path(args.xlsx))
    questions = build_questions(header, data, col_to_meta)

    write_sql(questions, Path(args.out_sql))
    write_json(questions, Path(args.out_json))

    # Summary stats
    by_section = Counter(q["section"] for q in questions)
    by_type = Counter(q["type"] for q in questions)
    by_tier = Counter(q["tier"] for q in questions)
    by_pathway = Counter(q["pathway"] for q in questions)

    print(f"Built {len(questions)} question metadata records\n")
    print("By section:")
    for s, n in sorted(by_section.items(), key=lambda x: -x[1]):
        print(f"  {n:4d}  {s}")
    print("\nBy type:")
    for t, n in sorted(by_type.items(), key=lambda x: -x[1]):
        print(f"  {n:4d}  {t}")
    print("\nBy tier:")
    for t in sorted(by_tier.keys()):
        print(f"  {by_tier[t]:4d}  tier {t}")
    print("\nBy pathway:")
    for p, n in sorted(by_pathway.items(), key=lambda x: -x[1]):
        print(f"  {n:4d}  {p}")
    print(f"\nOK: Wrote {args.out_sql}")
    print(f"OK: Wrote {args.out_json}")


if __name__ == "__main__":
    main()
