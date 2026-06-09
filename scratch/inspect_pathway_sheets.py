from openpyxl import load_workbook
wb = load_workbook('data/responses.xlsx', read_only=True)
for sheet in ['Intact Pathway', 'Circumcised Pathway', 'Meta Labels', 'Transformed Responses']:
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        # Get dimensions
        # Since read_only is True, we have to iterate to find size
        rows = list(ws.iter_rows(values_only=True))
        print(f"Sheet '{sheet}': Rows = {len(rows)}, Cols = {len(rows[0]) if rows else 0}")
        # Print first 5 rows
        for r in rows[:5]:
            if any(r):
                print("  ", [str(c)[:30] if c is not None else '' for c in r[:10]])
    else:
        print(f"Sheet '{sheet}' not found")
