from openpyxl import load_workbook
wb = load_workbook('data/responses.xlsx', read_only=True)
print("Sheet names:", wb.sheetnames)
