from openpyxl import load_workbook
wb = load_workbook('data/responses.xlsx', data_only=True)
ws = wb['Meta Tag Mapping']
rows = list(ws.iter_rows(values_only=True))
print("Meta Tag Mapping: Rows =", len(rows), "Cols =", len(rows[0]) if rows else 0)
for row in rows[:20]:
    if any(row):
        print(row)
