#!/usr/bin/env python3
"""
CircumSurvey — D1 data loader (v2, meta-tag aware)

Reads the raw Google Form responses xlsx (or CSV) and generates SQL INSERT
statements that can be executed against the D1 database via wrangler.

Usage:
  # Generate seed.sql from the xlsx downloaded from Drive:
  python3 scripts/load_d1.py --xlsx data/responses.xlsx --out data/seed.sql

  # Or from a CSV export:
  python3 scripts/load_d1.py --csv data/responses.csv --out data/seed.sql

  # Then push to remote D1:
  cd worker
  npx wrangler d1 execute circumsurvey --remote --file ../data/seed.sql

Key change from v1:
  Uses the "Meta Tag Mapping" sheet from the xlsx to get stable meta_tag
  slugs (demo_country_born, family_politics, etc.) instead of generating
  q### fallbacks. 355/373 columns get proper slugs; the remaining 18 are
  pathway-specific deep questions that still get stable q### slugs so
  ALL data is captured.
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

DEMO_DB_MAP = {
    "demo_country_born": "country_born",
    "demo_country_current": "country_now",
    "demo_us_state_born": "us_state_born",
    "demo_us_state_current": "us_state_now",
    "demo_can_province_born": "canada_province_born",
    "demo_can_province_current": "canada_province_now",
    "demo_ethnicity": "race_ethnicity",
    "demo_age": "age_bracket",
    "demo_generation": "generation",
    "demo_education_self": "education",
    "family_upbringing_status": "family_upbringing",
    "family_mother_education": "mother_education",
    "family_mother_profession": "mother_profession",
    "family_father_education": "father_education",
    "family_father_profession": "father_profession",
    "family_ses": "socioeconomic",
    "family_politics": "politics",
    "demo_sexuality": "sexuality",
    "demo_gender_identity": "gender",
    "demo_sex_assigned_at_birth": "sex_assigned",
}

RELIGION_DB_MAP = {
    "religion_is_significant": "upbringing_significance",
    "religion_primary_tradition": "primary_tradition",
    "religion_christian_denomination": "christian_denomination",
    "religion_jewish_denomination": "jewish_denomination",
    "religion_islamic_madhhab": "islamic_madhhab",
}

PATHWAY_SLUG = "pathway_state"

PATHWAY_NORMALIZATION = [
    (re.compile(r"^\s*intact", re.I), "intact"),
    (re.compile(r"^\s*circumcised\s*$", re.I), "circumcised"),
    (re.compile(r"restoring|restored", re.I), "restoring"),
    (re.compile(r"observer|N/A", re.I), "observer"),
]


def normalize_header(s):
    if not s:
        return ""
    s = str(s).strip()
    s = re.sub(r'[\U0001F300-\U0001FAFF\u2600-\u27BF\u2B00-\u2BFF\uFE0F\u20E3]+', '', s)
    s = ' '.join(s.split())
    s = s.strip(' \t\n\r.:;-–—')
    return s


def sqlstr(v):
    if v is None or v == "":
        return "NULL"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def clean_value(v):
    if v is None:
        return None
    v = str(v).strip()
    return v if v else None


def derive_pathway(val):
    if not val:
        return None
    for pat, label in PATHWAY_NORMALIZATION:
        if pat.search(str(val).strip()):
            return label
    return None


def extract_numeric(val):
    if val is None or val == "":
        return None
    s = str(val).strip()
    m = re.match(r"^\s*([1-5])(\s*[-–—:]|$)", s)
    if m:
        try: return float(m.group(1))
        except ValueError: return None
    try:
        n = float(s)
        if 1 <= n <= 5:
            return n
    except ValueError:
        pass
    return None


def load_from_xlsx(xlsx_path):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)

    header_to_meta = {}
    if "Meta Tag Mapping" in wb.sheetnames:
        ws_meta = wb["Meta Tag Mapping"]
        for row in ws_meta.iter_rows(values_only=True):
            if row and row[0] and row[1]:
                h = normalize_header(row[0])
                m = str(row[1]).strip()
                if h and m:
                    header_to_meta[h] = m
        header_to_meta_lower = {k.lower(): v for k, v in header_to_meta.items()}
    else:
        header_to_meta_lower = {}

    ws = wb["Form Responses 1"]
    rows = list(ws.iter_rows(values_only=True))
    form_header = rows[0]
    data_rows_raw = [r for r in rows[1:] if r and r[0]]

    col_to_meta = {}
    for i, h in enumerate(form_header):
        if not h:
            continue
        h_norm = normalize_header(h)
        if h_norm in header_to_meta:
            col_to_meta[i] = header_to_meta[h_norm]
        elif h_norm.lower() in header_to_meta_lower:
            col_to_meta[i] = header_to_meta_lower[h_norm.lower()]
        else:
            prefix = h_norm[:40].lower()
            for mh, ms in header_to_meta_lower.items():
                if prefix and (mh.startswith(prefix) or prefix in mh):
                    col_to_meta[i] = ms
                    break

    religion_manual = {
        24: "religion_christian_denomination", 25: "religion_christian_circ_view",
        26: "religion_christian_theology_basis", 27: "religion_christian_comments",
        28: "religion_jewish_denomination", 29: "religion_jewish_brit_milah_view",
        30: "religion_jewish_theology_awareness", 31: "religion_jewish_theology_reasons",
        32: "religion_jewish_identity_importance", 33: "religion_jewish_alt_interpretations",
        34: "religion_jewish_alt_thoughts", 35: "religion_jewish_diversity_room",
        36: "religion_jewish_brit_shalom", 37: "religion_islamic_madhhab",
        38: "religion_islamic_khitan_view", 39: "religion_islamic_religious_awareness",
        40: "religion_islamic_religious_reasons", 41: "religion_islamic_fard_or_sunnah",
        42: "religion_islamic_identity_importance", 43: "religion_islamic_alt_interpretations",
        44: "religion_islamic_alt_thoughts", 45: "religion_islamic_diversity_room",
        46: "religion_islamic_intactness_reconcile",
    }
    for k, v in religion_manual.items():
        if k not in col_to_meta:
            col_to_meta[k] = v

    col_to_meta[65] = PATHWAY_SLUG

    for i in range(len(form_header)):
        if form_header[i] is not None and i not in col_to_meta:
            col_to_meta[i] = f"q{i:03d}"

    data_rows = [list(r) + [None] * (len(form_header) - len(r)) for r in data_rows_raw]
    return form_header, data_rows, col_to_meta


def load_from_csv(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        form_header = next(reader)
        rows = [r for r in reader if r and r[0]]
    data_rows = [r + [""] * (len(form_header) - len(r)) for r in rows]
    col_to_meta = {i: f"q{i:03d}" for i in range(len(form_header)) if form_header[i]}
    col_to_meta[65] = PATHWAY_SLUG
    return form_header, data_rows, col_to_meta


def generate_sql(form_header, data_rows, col_to_meta, out_path):
    slug_to_col = {}
    for col_idx, slug in col_to_meta.items():
        if slug not in slug_to_col:
            slug_to_col[slug] = col_idx

    sql = []
    sql.append("-- ═══════════════════════════════════════════════════════════════")
    sql.append("-- CircumSurvey — Seed data generated by scripts/load_d1.py v2")
    sql.append(f"-- Rows: {len(data_rows)}")
    sql.append("-- Run: npx wrangler d1 execute circumsurvey --remote --file ../data/seed.sql")
    sql.append("-- ═══════════════════════════════════════════════════════════════")
    sql.append("")
    # Note: D1 imports via `wrangler d1 execute --file` handle
    # transactions internally. Per Cloudflare docs, DO NOT wrap
    # the file in BEGIN TRANSACTION / COMMIT or wrangler errors with
    # "cannot start a transaction within a transaction".
    sql.append("PRAGMA defer_foreign_keys = true;")
    sql.append("DELETE FROM responses;")
    sql.append("DELETE FROM religion;")
    sql.append("DELETE FROM demographics;")
    sql.append("DELETE FROM respondents;")
    sql.append("")

    pathway_col = slug_to_col.get(PATHWAY_SLUG, 65)
    valid = []
    for idx, row in enumerate(data_rows):
        ts = clean_value(row[0]) if row else None
        if not ts:
            continue
        pathway = derive_pathway(row[pathway_col] if pathway_col < len(row) else None)
        valid.append((len(valid) + 1, idx, row, pathway, ts))

    for rid, src, _, pathway, ts in valid:
        sql.append(
            f"INSERT INTO respondents (id, submitted_at, pathway, consent, source_row_idx) "
            f"VALUES ({rid}, {sqlstr(ts)}, {sqlstr(pathway)}, 1, {src});"
        )

    sql.append("")
    sql.append("-- Demographics")
    for rid, _, row, _, _ in valid:
        cols = ["respondent_id"]
        vals = [str(rid)]
        for meta_slug, db_col in DEMO_DB_MAP.items():
            col_idx = slug_to_col.get(meta_slug)
            v = clean_value(row[col_idx]) if col_idx is not None and col_idx < len(row) else None
            cols.append(db_col)
            vals.append(sqlstr(v))
        sql.append(f"INSERT INTO demographics ({', '.join(cols)}) VALUES ({', '.join(vals)});")

    sql.append("")
    sql.append("-- Religion")
    for rid, _, row, _, _ in valid:
        cols = ["respondent_id"]
        vals = [str(rid)]
        for meta_slug, db_col in RELIGION_DB_MAP.items():
            col_idx = slug_to_col.get(meta_slug)
            v = clean_value(row[col_idx]) if col_idx is not None and col_idx < len(row) else None
            cols.append(db_col)
            vals.append(sqlstr(v))
        sql.append(f"INSERT INTO religion ({', '.join(cols)}) VALUES ({', '.join(vals)});")

    sql.append("")
    sql.append("-- Responses (long-form, every non-empty cell)")
    resp_count = 0
    for rid, _, row, _, _ in valid:
        for col_idx in range(len(row)):
            val = clean_value(row[col_idx])
            if val is None:
                continue
            slug = col_to_meta.get(col_idx)
            if not slug:
                continue
            num = extract_numeric(val)
            if num is not None:
                sql.append(
                    f"INSERT OR REPLACE INTO responses (respondent_id, question_id, value_text, value_num) "
                    f"VALUES ({rid}, {sqlstr(slug)}, {sqlstr(val)}, {num});"
                )
            else:
                sql.append(
                    f"INSERT OR REPLACE INTO responses (respondent_id, question_id, value_text) "
                    f"VALUES ({rid}, {sqlstr(slug)}, {sqlstr(val)});"
                )
            resp_count += 1

    sql.append("")
    sql.append(f"-- {len(valid)} respondents, {resp_count} response rows")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(sql), encoding="utf-8")
    print(f"Wrote {out_path}")
    print(f"  {len(valid)} respondents, {resp_count} response rows")

    from collections import Counter
    dist = Counter(p for _, _, _, p, _ in valid)
    print(f"\n  Pathway distribution:")
    for k in ["intact", "circumcised", "restoring", "observer", None]:
        print(f"    {k or '(unclassified)':15s}  {dist.get(k, 0):4d}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", help="Path to responses.xlsx from Drive")
    ap.add_argument("--csv", help="Path to raw Form responses CSV")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    if args.xlsx:
        header, rows, mapping = load_from_xlsx(Path(args.xlsx))
    elif args.csv:
        header, rows, mapping = load_from_csv(Path(args.csv))
    else:
        print("ERROR: Must provide --xlsx or --csv", file=sys.stderr)
        sys.exit(1)

    q_count = sum(1 for v in mapping.values() if v.startswith('q') and v[1:].isdigit())
    print(f"Loaded {len(header)} columns, {len(rows)} rows")
    print(f"  {len(mapping) - q_count} meaningful slugs, {q_count} q### fallbacks")
    generate_sql(header, rows, mapping, Path(args.out))
