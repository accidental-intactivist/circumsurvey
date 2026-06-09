from openpyxl import load_workbook
wb = load_workbook('data/responses.xlsx', data_only=True)
if 'Meta Labels' in wb.sheetnames:
    ws = wb['Meta Labels']
    print("=== Meta Labels Sheet ===")
    for row in list(ws.iter_rows(values_only=True))[:50]:
        if any(row):
            print([str(cell)[:40] if cell is not None else '' for cell in row])
else:
    print("Meta Labels not found")
