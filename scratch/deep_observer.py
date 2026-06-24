import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(sys.argv[1], data_only=True)

# Check the Meta Tag Mapping for pathway clues
if "Meta Tag Mapping" in wb.sheetnames:
    ws_meta = wb["Meta Tag Mapping"]
    print("=== Meta Tag Mapping entries containing 'path' or 'observer' ===")
    for row in ws_meta.iter_rows(values_only=True):
        if row and row[0] and row[1]:
            h = str(row[0]).strip()
            m = str(row[1]).strip()
            if any(w in h.lower() + m.lower() for w in ['pathway', 'observer', 'path_state', 'which best']):
                print(f"  {m:40s} <- {h[:80]}")

ws = wb['Form Responses 1']
rows = list(ws.iter_rows(values_only=True))
header = rows[0]

# Find the pathway column by searching headers
print("\n=== Headers containing 'path', 'status', 'which best' ===")
for i, h in enumerate(header):
    if h and any(w in str(h).lower() for w in ['pathway', 'which best describes', 'circumcision status']):
        print(f"  col {i}: {str(h)[:120]}")

# Check what column 65 actually is
print(f"\n=== Column 65 header ===")
print(f"  {str(header[65])[:120] if len(header) > 65 and header[65] else 'NONE/EMPTY'}")

# For empty-pathway rows, just show first 20 columns
print("\n=== First empty-pathway row: first 20 non-empty columns ===")
count = 0
for row in rows[1:]:
    if len(row) > 65 and row[65] is None:
        count += 1
        if count > 1:
            continue
        for i in range(min(20, len(row))):
            if row[i] is not None and str(row[i]).strip():
                h = str(header[i])[:60] if i < len(header) and header[i] else f"col{i}"
                print(f"  col {i}: {str(row[i])[:80]}  (hdr: {h})")

print(f"\nTotal empty-pathway rows: {count}")
